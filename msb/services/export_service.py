from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


@dataclass
class BadgeInfo:
    participant_id: int
    full_name: str
    job: str
    is_guest: bool
    tables: list[str]


class ExportService:
    """Service d'export (Excel, PDF Badges)."""

    def __init__(self, persistence=None, logo_path: Path | None = None) -> None:
        self.persistence = persistence
        self.logo_path = logo_path

    def export_excel(self, output_path: str | Path) -> Path:
        """Exporte le plan (vues tables + participants) au format Excel."""

        return self.export_plan_excel(output_path)

    def export_plan_excel(self, output_path: str | Path) -> Path:
        """Génère un Excel contenant le plan de table (vues table et participant)."""

        persistence = self._require_persistence()
        output_path = Path(output_path)

        event_info = persistence.get_event_info()
        participants = list(persistence.list_participants())
        plan = persistence.load_plan()

        if not plan:
            raise RuntimeError("Aucun plan de table enregistré. Générez ou chargez un plan avant d'exporter.")

        participants_by_id = {p.id: p for p in participants}
        session_count = len(plan)
        table_count = len(plan[0]) if session_count else 0

        wb = Workbook()

        # Vue par table
        ws_tables = wb.active
        ws_tables.title = "Plan par table"
        ws_tables.append(["Session / Table", *[f"Table {i + 1}" for i in range(table_count)]])

        for s_idx, tables in enumerate(plan):
            row = [f"Session {s_idx + 1}"]
            for t_idx, pids in enumerate(tables):
                names = []
                for pid in pids:
                    p = participants_by_id.get(pid)
                    if p:
                        names.append(f"{p.first_name} {p.last_name} - {p.job}")
                row.append("\n".join(names) if names else "-")
            ws_tables.append(row)

        wrap_align = Alignment(wrap_text=True, vertical="top")
        for row in ws_tables.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.alignment = wrap_align

        # Vue par participant
        ws_by_participant = wb.create_sheet("Plan par participant")
        ws_by_participant.append(["Participant", *[f"S{i + 1}" for i in range(session_count)]])

        for p in participants:
            row = [f"{p.first_name} {p.last_name} ({p.job})"]
            for s_idx in range(session_count):
                table_idx = "-"
                tables = plan[s_idx]
                for t_idx, pids in enumerate(tables):
                    if p.id in pids:
                        table_idx = t_idx + 1
                        break
                row.append(table_idx)
            ws_by_participant.append(row)

        ws_by_participant.freeze_panes = "B2"
        ws_tables.freeze_panes = "B2"

        # Résumé minimal
        summary = wb.create_sheet("Résumé")
        summary.append(["Événement", event_info.get("name", "")])
        summary.append(["Sessions", session_count])
        summary.append(["Tables", table_count])
        summary.append(["Participants", len(participants)])

        wb.save(output_path)
        return output_path

    def export_import_template(self, output_path: str | Path) -> Path:
        """Exporte un modèle Excel pour réimport de participants."""

        persistence = self._require_persistence()
        output_path = Path(output_path)

        wb = Workbook()
        ws = wb.active
        ws.title = "Participants"

        headers = [
            "Prénom",
            "Nom",
            "Métier",
            "Visiteur (Oui/Non)",
            "Chef de table (Oui/Non)",
        ]
        ws.append(headers)

        participants = list(persistence.list_participants())
        if participants:
            for p in participants:
                ws.append(
                    [
                        p.first_name,
                        p.last_name,
                        p.job,
                        "Oui" if getattr(p, "is_guest", False) else "Non",
                        "Oui" if getattr(p, "is_table_lead", False) else "Non",
                    ]
                )
        else:
            ws.append(["Alice", "Durand", "Conseillère financière", "Non", "Oui"])
            ws.append(["Bob", "Martin", "Architecte", "Oui", "Non"])

        wb.save(output_path)
        return output_path

    def export_badges_pdf(self, output_path: str | Path) -> Path:
        """
        Génère un PDF contenant un badge par participant.

        Format demandé :
        - Nom de la réunion + logo MySpeedBusiness
        - Nom & prénom, métier, suffixe « (Invité) » si applicable
        - Table assignée pour chaque session, dans l'ordre
        """
        persistence = self._require_persistence()
        output_path = Path(output_path)

        event_info = persistence.get_event_info()
        participants = list(persistence.list_participants())
        plan = persistence.load_plan()

        session_count = event_info.get("session_count") or 0
        session_count = max(session_count, len(plan)) if plan else session_count

        badges = self._build_badges(participants, plan, session_count)
        logo = self._resolve_logo_path()

        self._render_badges(
            output_path=output_path,
            event_name=event_info.get("name", ""),
            badges=badges,
            logo_path=logo,
        )
        return output_path

    # --- helpers ---------------------------------------------------------
    def _require_persistence(self):
        if not self.persistence:
            raise RuntimeError("Persistence non fournie pour l'export")
        return self.persistence

    def _resolve_logo_path(self) -> Path | None:
        if self.logo_path:
            return Path(self.logo_path)
        default = Path(__file__).resolve().parents[2] / "img" / "msb_logo.png"
        return default if default.exists() else None

    def _build_badges(self, participants: Iterable, plan: list, session_count: int) -> list[BadgeInfo]:
        tables_by_participant = {
            p.id: ["-"] * session_count for p in participants
        } if session_count > 0 else {p.id: [] for p in participants}

        for s_idx, tables in enumerate(plan or []):
            if s_idx >= session_count:
                break
            for t_idx, pids in enumerate(tables):
                for pid in pids:
                    if pid in tables_by_participant:
                        tables_by_participant[pid][s_idx] = str(t_idx + 1)

        badges: list[BadgeInfo] = []
        for p in participants:
            full_name = f"{p.first_name} {p.last_name}"
            badges.append(
                BadgeInfo(
                    participant_id=p.id,
                    full_name=full_name,
                    job=p.job,
                    is_guest=bool(getattr(p, "is_guest", False)),
                    tables=tables_by_participant.get(p.id, []),
                )
            )
        return badges

    def _render_badges(
        self,
        *,
        output_path: Path,
        event_name: str,
        badges: list[BadgeInfo],
        logo_path: Path | None,
    ) -> None:
        c = canvas.Canvas(str(output_path), pagesize=A4)
        page_width, page_height = A4

        badge_width = 90 * mm
        badge_height = 55 * mm
        margin = 10 * mm
        h_spacing = 5 * mm
        v_spacing = 5 * mm

        cols = max(1, int((page_width - 2 * margin + h_spacing) // (badge_width + h_spacing)))
        rows = max(1, int((page_height - 2 * margin + v_spacing) // (badge_height + v_spacing)))
        badges_per_page = max(1, cols * rows)

        logo_reader = None
        logo_max_width = 30 * mm
        logo_max_height = 12 * mm
        if logo_path and logo_path.exists():
            logo_reader = ImageReader(str(logo_path))

        palette = [
            colors.HexColor("#ffd84a"),
            colors.HexColor("#f06292"),
            colors.HexColor("#4fc3f7"),
            colors.HexColor("#aed581"),
            colors.HexColor("#ff8a65"),
            colors.HexColor("#f48fb1"),
            colors.HexColor("#29b6f6"),
            colors.HexColor("#7986cb"),
        ]

        for idx, badge in enumerate(badges):
            pos_in_page = idx % badges_per_page
            if idx and pos_in_page == 0:
                c.showPage()

            col = pos_in_page % cols
            row = pos_in_page // cols

            x = margin + col * (badge_width + h_spacing)
            y = page_height - margin - (row + 1) * badge_height - row * v_spacing

            self._draw_badge(
                c=c,
                origin_x=x,
                origin_y=y,
                width=badge_width,
                height=badge_height,
                badge=badge,
                event_name=event_name,
                palette=palette,
                logo_reader=logo_reader,
                logo_size=(logo_max_width, logo_max_height),
            )

        c.save()

    def _draw_badge(
        self,
        *,
        c: canvas.Canvas,
        origin_x: float,
        origin_y: float,
        width: float,
        height: float,
        badge: BadgeInfo,
        event_name: str,
        palette: list,
        logo_reader: ImageReader | None,
        logo_size: tuple[float, float],
    ) -> None:
        padding = 6 * mm
        box_gap = 2 * mm
        strip_height = 16 * mm
        bottom_bar_height = 10 * mm

        c.saveState()
        c.translate(origin_x, origin_y)

        # Contour du badge
        c.setLineWidth(1)
        c.roundRect(0, 0, width, height, radius=4 * mm, stroke=1, fill=0)

        y = height - padding

        # En-tête de réunion
        c.setFillColor(colors.HexColor("#c62828"))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(padding, y - 8, event_name or "")

        if logo_reader:
            logo_w, logo_h = logo_size
            c.drawImage(
                logo_reader,
                width - padding - logo_w,
                height - padding - logo_h,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                mask="auto",
            )

        # Nom complet
        y -= 22
        name_line = badge.full_name
        if badge.is_guest:
            name_line = f"{name_line} (Invité)"
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(padding, y, name_line)

        # Métier
        y -= 14
        c.setFont("Helvetica", 11)
        c.drawString(padding, y, badge.job)

        # Bandeau sessions / tables
        y -= 18
        c.setFont("Helvetica-Bold", 9)
        c.drawString(padding, y, "Sessions / Tables")

        if badge.tables:
            count = len(badge.tables)
            avail_width = width - 2 * padding - (count - 1) * box_gap
            box_width = avail_width / count if count else 0
            box_height = strip_height
            base_y = y - box_height - 4
            for idx, table in enumerate(badge.tables):
                color = palette[idx % len(palette)] if palette else colors.lightgrey
                x = padding + idx * (box_width + box_gap)
                c.setFillColor(color)
                c.roundRect(x, base_y, box_width, box_height, radius=2 * mm, stroke=0, fill=1)

                c.setFillColor(colors.black)
                c.setFont("Helvetica-Bold", 10)
                c.drawCentredString(x + box_width / 2, base_y + box_height - 8, f"S{idx + 1}")
                c.setFont("Helvetica", 9)
                c.drawCentredString(x + box_width / 2, base_y + 8, f"Table {table}")
            y = base_y - 6
        else:
            y -= 10
            c.setFont("Helvetica", 9)
            c.drawString(padding, y, "Aucun plan de table enregistré.")

        # Bandeau d'état (membre / invité)
        c.setFillColor(colors.HexColor("#b86d1f"))
        c.rect(0, 0, width, bottom_bar_height, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 12)
        status = "Invité" if badge.is_guest else "Membre"
        c.drawCentredString(width / 2, bottom_bar_height / 2 - 3, status)

        c.restoreState()
