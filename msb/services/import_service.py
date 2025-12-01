from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook


class ImportService:
    """Service d'import (Excel & UI)."""

    def __init__(self, persistence=None) -> None:
        self.persistence = persistence

    # --- public API ------------------------------------------------------
    def import_from_excel(self, file_path: str | Path) -> int:
        """Importe les participants depuis un fichier Excel.

        Le format attendu est une feuille avec les colonnes :
        - Prénom
        - Nom
        - Métier
        - Visiteur (Oui/Non)
        - Chef de table (Oui/Non)
        """

        persistence = self._require_persistence()
        wb = load_workbook(filename=file_path)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0

        header = [self._normalize_header(h) for h in rows[0]]
        col_idx = self._map_columns(header)

        added = 0
        for raw in rows[1:]:
            if not raw or all(v is None or str(v).strip() == "" for v in raw):
                continue

            first = self._read_cell(raw, col_idx.get("first_name"))
            last = self._read_cell(raw, col_idx.get("last_name"))
            job = self._read_cell(raw, col_idx.get("job"))
            guest = self._parse_bool(raw[col_idx["is_guest"]]) if col_idx.get("is_guest") is not None else False
            lead = (
                self._parse_bool(raw[col_idx["is_table_lead"]])
                if col_idx.get("is_table_lead") is not None
                else False
            )

            if not first or not last or not job:
                # ligne incomplète : on ignore
                continue

            persistence.add_participant(first, last, job, guest, lead)
            added += 1

        return added

    def import_from_ui(self, rows: list[dict]) -> int:
        """Importe des participants fournis par l'UI.

        Chaque dict doit contenir first_name, last_name, job et, optionnellement,
        is_guest / is_table_lead.
        """

        persistence = self._require_persistence()
        added = 0

        for row in rows:
            first = str(row.get("first_name", "")).strip()
            last = str(row.get("last_name", "")).strip()
            job = str(row.get("job", "")).strip()
            if not first or not last or not job:
                continue
            is_guest = bool(row.get("is_guest", False))
            is_lead = bool(row.get("is_table_lead", False))
            persistence.add_participant(first, last, job, is_guest, is_lead)
            added += 1

        return added

    # --- helpers ---------------------------------------------------------
    def _require_persistence(self):
        if not self.persistence:
            raise RuntimeError("Persistence non fournie pour l'import")
        return self.persistence

    def _normalize_header(self, value) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
        # Ignore additional hints such as "(oui/non)" that may appear in the header
        if "(" in text:
            text = text.split("(", 1)[0].strip()
        return text

    def _map_columns(self, header: list[str]) -> dict[str, int]:
        mapping = {
            "first_name": {"prenom", "first_name"},
            "last_name": {"nom", "last_name"},
            "job": {"metier", "job", "fonction"},
            "is_guest": {"visiteur", "guest", "invite"},
            "is_table_lead": {"chefdetable", "chef de table", "tablelead", "leader", "chef"},
        }

        idx = {key: None for key in mapping}
        for i, col in enumerate(header):
            for field, names in mapping.items():
                if col.replace(" ", "") in names:
                    idx[field] = i
        if idx["first_name"] is None or idx["last_name"] is None or idx["job"] is None:
            raise ValueError("Colonnes obligatoires manquantes dans l'Excel")
        return idx

    def _parse_bool(self, value) -> bool:
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().casefold()
        truthy = {"oui", "yes", "true", "1", "y", "o"}
        falsy = {"non", "no", "false", "0", "n"}
        if text in truthy:
            return True
        if text in falsy:
            return False
        return False

    def _read_cell(self, row: tuple, index: int | None) -> str:
        if index is None:
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()
